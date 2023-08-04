# -*- coding: utf-8 -*-
"""
Created on Sun Jul 30 13:17:10 2023

@author: Matt0
"""
import pandas as pd
import numpy as np
from openpyxl.workbook import workbook
from openpyxl import load_workbook
from scipy.stats import chi2_contingency
import seaborn as sns
import matplotlib.pyplot as plt
from scipy import stats
import statsmodels.api as sm
from statsmodels.compat import lzip
import statsmodels.formula.api as smf
import statsmodels.stats.api as sms
import Scorers as scorer
import Risk_assign as risk
from tabulate import tabulate
from pathlib import Path


current_dir=Path.cwd()
file_names=["Script.xlsx","Framingham practice more.txt"]

script_path=current_dir / file_names[0]
data_path= current_dir / file_names[1]


wb=load_workbook(script_path)
ws=wb.active

def enter_new():
    print("Please enter some information so I can calculate the scores:\n")

    name=str(input("What is the person's name?  "))
    
    sex=str(input("What is the person's sex? M for male, F for female  ")).upper()
    while sex.upper() not in ["M","F"]:
        sex=str(input("\n You must enter either M or F here, as these scores are not validated in other groups. Use assigned sex at birth here.  ")).upper()
    
    Cholesterol=int(input("What is the person's cholesterol level?  "))
    if (100>=Cholesterol) or (Cholesterol>=300):
        Cholesterol=int(input("\n Typical cholesterol levels don't go below 100 or above 300, reenter the number to confirm.  "))
    
    HDL=int(input("What is their HDL cholesterol level?  "))
    if (20>=HDL) or (HDL>=100):
        HDL=int(input("\n Typical HDL cholesterol levels don't go below 20 or above 100, reenter the number to confirm.  "))
        
    Systolic=int(input("What is their systolic blood pressure?  "))
    if (70>=Systolic) or (Systolic>=180):
        Systolic=int(input("\n Looking at these numbers, the person might be in serious distress. Please get them help and reenter to confirm or fix the entry. "))
    
    Treated=input("Are they being treated for high blood pressure medications? Enter yes or no  ").capitalize()
    while Treated.capitalize() not in ["Yes","No"]:
        Treated=str(input("\n You must enter either yes or no here.  ")).capitalize()
        
    Smoking=str(input("Are they currently a smoker? Enter yes or no.  ")).capitalize()
    while Smoking.capitalize() not in ["Yes","No"]:
        Smoking=str(input("\n You must enter either yes or no here.  ")).capitalize()
    
    Age=int(input("What is their age?  "))
    if (Age<30) or (Age>80):
        Age=int(input("These scores are not validated for people below 70 or older than 30. To continue, enter either 30 or 70 in place for a rough estimate.  "))
        
    NewPat=pd.DataFrame([[name,sex,Cholesterol,HDL,Systolic,Treated,Smoking, Age]],columns=['Name','Gender', 'Cholest', "HDL", "Systolic", "Treat","Smoke","Age"])
    
    
    NewPat, YourRaws=calc_all(NewPat)
    
    NewPat=pd.concat([NewPat,risk.risk_assign(NewPat['Framingham'],NewPat['Gender'])],axis=1)


    print("\n\n *****************")
    print(f"Name: {NewPat['Name'][0]}")
    print(f"Sex: {NewPat['Gender'][0]}\n")
    print(tabulate([["Indicator","Original", "Scored"],
                    ["Age", float(NewPat['Age'].iloc[0]), float(YourRaws['age_scr'].iloc[0])],
                    ["Total Cholesterol" ,float(NewPat['Cholest'].iloc[0]), float(YourRaws['cholest_scr'].iloc[0])],
                    ["HDL Cholesterol",float(NewPat['HDL'].iloc[0]), float(YourRaws['hdl_scr'].iloc[0])],
                    ["Smoker Status", NewPat['Smoke'].iloc[0],float(YourRaws['smoke_scr'].iloc[0])],
                    ["Systolic BP", float(NewPat['Systolic'].iloc[0]),float(YourRaws['systol_scr'].iloc[0])],
                     ],headers="firstrow"))
    print("\n")
    print(tabulate([["Framingham Score","10-year Risk of CVD"],
                    [NewPat['Framingham'].loc[0],f"{NewPat['Risk_pct'].loc[0]}%"]],
                   headers="firstrow"))
def calc_all(Patients):

    rawscores=pd.DataFrame()
    rawscores['age_scr']=scorer.age_scorer(Patients)
    rawscores['cholest_scr']=scorer.cholest_scorer(Patients)
    rawscores['hdl_scr']=scorer.hdl_scorer(Patients)
    
    rawscores['smoke_scr']=scorer.smoke_scorer(Patients)
    rawscores['systol_scr']=scorer.systolic_scorer(Patients)
    
    Patients['Framingham']=rawscores.sum(axis=1)
    return Patients, rawscores

def demonstrate():
    Patients=pd.read_csv(r"C:\Users\Matt0\test project\Framingham score tool\Framingham practice more.txt").rename(
        columns={'Total_Cholesterol':'Cholest','HDL_Cholesterol':'HDL',
                 'Smoking_Status':'Smoke', 'Systolic_Blood_Pressure':'Systolic',
                 'Treatment_for_High_BP':'Treat'})
    #purge missings if any
    
    #introduce the data
    print("\n\n_____________________________________________\n\n")

    print(ws["C2"].value)
    
    print("\n\n_____________________________________________\n\n")
    
    print(Patients.head(5))
    input("\n\nPress enter to start calculating the Framingham scores")
    
    Patients,rawscores=calc_all(Patients)
    
    print("\n\n_____________________________________________\n\n")
    
    #introduce the score set
    print(f"""{ws["C3"].value}\n\n""")
    
    print(rawscores.head(5))
    
    input("\n\nPress enter to continue")
    
    print("\n\n_____________________________________________\n\n")
    print("\n\n The rows are then combined and merged into the main dataset\n\n")
    print(Patients.head(5))
    
    input("\n\nPress enter to continue")
    
    ###compute risk tiers
    
    Patients=pd.concat([Patients,risk.risk_assign(Patients['Framingham'],Patients['Gender'])],axis=1)
    
    print("\n\n_____________________________________________\n\n")
    
    
    print(f"""\n\n{ws["C4"].value}\n\n""")
    
    print(Patients[['Gender', 'Framingham','Risk_pct','Riskcats','Highrisk']].head(5))
    
    print("\n\nThats it! Now the numbers can be used for analysis.")
    input("\n\n ******Press enter to go back to the main hub.******")

#pd.DataFrame.to_csv(Patients,"almost prepared data")    

def gendata(Patients):
    np.random.seed(29293)
    Patients['tempcost']=np.random.randint(85,10000, 250)
    #Impose a correlation between this new variable and Framingham scores
    Patients['Annual_med_costs']=(Patients['Framingham']+.002*Patients['tempcost'])*100
    del Patients['tempcost']


    #An indicator of if they are a snap recipient or not
    Patients['SNAP']=np.random.default_rng().choice(["no","yes"],len(Patients['Gender']))
    #doctor it so Patients is correlated with high risk by randomly assigning some of the snap people to be high risk.
    rand=np.random.randint(0,100,250)
    Patients.loc[(rand>=35) & (Patients['Highrisk']==">20% risk"), 'SNAP']="yes"

    #dummy up gender and SNAP for later
    Patients['Male']=pd.get_dummies(Patients['Gender'],drop_first=True, dtype=float)
    Patients['SNAP_yes']=pd.get_dummies(Patients['SNAP'], drop_first=True, dtype=float)

    return Patients


def show_risky(Patients):
    ########STEP 1: Ask if they want to see high risk patients.
    #Show if yes, move on if no
    #
    #get the IDs for people
    atrisk=pd.DataFrame(Patients.loc[Patients['Highrisk']=='>20% risk','Patient_ID'])
    atrisk=pd.merge(atrisk,Patients[['Patient_ID','Risk_pct']], on='Patient_ID')
    
    ##sns.histogram(Patients, x='Framingham')
    text=ws['E2'].value
    text= text.replace("{len(atrisk)}",f'{len(atrisk)}')
    input(text)
    print("\n\n_____________________________________________\n\n")
    header="The following patients IDs have above 20% estimated risk of coronary heart disease in the next decade. \n\n Below is each ID and risk % \n"
    header=header.center(45)
    print(header)
    for case in atrisk.itertuples():
        print(f'Patient ID: {case[1]}, risk of {case[2]}%')
            
    input("\n\n ******Press enter to go back to the main hub.******")

def crosstab(Patients):
### A cross tab and chi square:
        
    input(ws['G2'].value)

    
    print("\n\n_____________________________________________\n\n")  
    print("Here's a cross tab of risk scores and SNAP values, calculating percentages along columns. \n\n")
    nopct=pd.crosstab(Patients['Highrisk'],Patients['SNAP'])
    chi, p, df, expected= chi2_contingency(nopct)
    tab=(pd.crosstab(Patients['Highrisk'],Patients['SNAP'],normalize='columns').round(4)*100)
    cramersv=np.sqrt((chi/len(Patients['SNAP']*(len(tab)-1))))
    print(tab)
    
    print(f"\nChi-Square= {chi.round(2)}, p= {p.round(3)}")
    print(f"Cramer's V= {cramersv.round(3)}")

    
    explain=str(input("\n Shall I provide an interpretation of these results and next steps? Y for yes, N to go back to the menu\n"))
    while explain.lower() not in ['y','n']:
        
        explain=str(input("Please enter either Y or N \n"))

    if explain.lower()=="y":
        print("\n\n_____________________________________________\n\n")
        print(ws['G3'].value)
        input("\n\n ******Press enter to go back to the main hub.******")

def makehistogram(Patients):
        
    input(ws['I2'].value)

    print("\n\n_____________________________________________\n\n")  
    print("First, it is important to examine distributions for both variables.\n\n")
      
    plt.rcParams["figure.figsize"] = [7.00, 3.50]
    plt.rcParams["figure.autolayout"] = True
    f, axes = plt.subplots(1, 2)
    
    sns.histplot(data=Patients,
                 x='Risk_pct',
                 bins=20,
                 ax=axes[0])
    axes[0].set_xlabel("10-year risk %")
    axes[0].set_title("Framingham Risk Estimate Distribution")
    axes[0].axvline(x=Patients['Risk_pct'].mean(),color='red')
    axes[0].axvline(x=Patients['Risk_pct'].median(),color='blue')
    
    sns.histplot(Patients['Annual_med_costs'],
                 bins=20,
                 ax=axes[1])
    axes[1].set_title("Annual Medical Costs Distribution")
    axes[1].set_xlabel("$ Spent Per Year")
    axes[1].axvline(x=Patients['Annual_med_costs'].mean(),color='red')
    axes[1].axvline(x=Patients['Annual_med_costs'].median(),color='blue')
    
    plt.show()
    
    explain=str(input("\n Would you like an explanation of these graphs? Enter Y or N?\n"))
    while explain.lower() not in ['y','n']:
        
        explain=str(input("Please enter either Y or N \n"))

    if explain.lower()=="y":
        print("\n\n_____________________________________________\n\n")
        explanation=ws['I3'].value
        mean=int(Patients['Annual_med_costs'].mean())
        std=int(Patients['Annual_med_costs'].std())
        stdrange=int(std*1.96)
        explanation=explanation.replace('PATIENTMEAN',f'${mean}')
        explanation=explanation.replace('PATIENTSTD',f'${std}')
        explanation=explanation.replace('STDRANGE',f'${stdrange}')
        input(explanation)
            
def runscatter(Patients):
    
###Scatter plots

###display the first
    input(ws['I4'].value)
    plt.figure(figsize=(8, 5))
    sns.regplot(x = "Risk_pct", y = "Annual_med_costs", data = Patients)
    plt.title("Correlation between Framingham scores \n and medical expenditures", fontsize=18)
    plt.xlabel("10-year risk %", fontsize=15)
    plt.ylabel("Annual medical expenditures", fontsize=15)
    plt.show()
    
    print("\n_____________________________________________\n")
    
    ##display the second
    input(ws['I5'].value)
    
    print("\n_____________________________________________\n")
    plt.figure(figsize=(8, 5))
    sns.lmplot(x = "Risk_pct", y = "Annual_med_costs",
               hue = "Gender", 
               markers = ["s", "x"],
               data = Patients,
               legend=False,
               facet_kws={'legend_out': True},
               height=5,
               aspect=1.5)
    plt.legend(title="Gender", loc="upper left", bbox_to_anchor=(.75, .25))
    plt.title("Highlighting sex differences", fontsize=18)
    plt.xlabel("10-year risk %", fontsize=15)
    plt.ylabel("Annual medical expenditures", fontsize=15)
    
    plt.show()
    interpretplz=input(ws['I6'].value)
    while interpretplz.lower() not in ['y','n']:
        interpretplz=str(input("Please enter either Y or N \n"))
    
    if interpretplz=="y":
        answer=ws['I7'].value
        answer=answer.replace("CORRCOEF",f"{Patients['Risk_pct'].corr(Patients['Annual_med_costs']).round(2)}")
        print("\n")
        input(answer)
        
        
        
def OLSdemo(Patients):
    ####Do a regression model and diagnostics
    
    # predictor = sm.add_constant(Patients['Risk_pct'])
    # model=sm.OLS(Patients['Annual_med_costs'],predictor).fit()
    # output=model.summary()
    # print(output)
    
    #more covs
    print("\n\n_____________________________________________\n\n")
    
    input(ws['I8'].value)
    
    model=smf.ols('Annual_med_costs~Risk_pct+Male+SNAP_yes',data=Patients)
    output=model.fit()
    print(output.summary())
    
    print("\n\n_____________________________________________\n\n")
    
    print("Above are the base model estimates.")
          
    
    options=input(ws['I9'].value)
    while options not in ['1','2','3','4']:
            options=input("You have to enter a value from 1 to 4")
    
           
    while options!='4':
        if options=='1':
        ##standardized coefs
            Patients['Annual_med_costs_Beta']=stats.zscore(Patients['Annual_med_costs'])
            
            Patients['Risk_pct_Beta']=stats.zscore(Patients['Risk_pct'])
            Patients['SNAP_yes_Beta']=stats.zscore(Patients['SNAP_yes'])
            Patients['Male_Beta']=stats.zscore(Patients['Male'])
            
            
            varliststd=Patients[['Risk_pct_Beta','SNAP_yes_Beta','Male_Beta']]
            varliststd = sm.add_constant(varliststd)
            varliststd_standardized = (varliststd - varliststd.mean()) / varliststd.std()
            
            # Fit the model
            modelstd = sm.OLS(Patients['Annual_med_costs_Beta'], varliststd).fit()
            outputstd=modelstd.summary()
            print(outputstd)
            
            input("\n\n*********Press enter to go back to the regression menu*********")
        
        if options=='2':
            print("\n\n_____________________________________________\n\n")
            input("We'll start with a residual vs fitted plots and other similar tools. Press enter to show.")
            diags = plt.figure(figsize=(12,8))
            sm.graphics.plot_regress_exog(output, 'Risk_pct', fig=diags)
            plt.show(diags)
            
            print("\n\n_____________________________________________\n\n")
    
            input("Press enter to show a graph of potential outliers using Cook's D. ")
            fig = sm.graphics.influence_plot(output, criterion="cooks")
            fig.tight_layout(pad=1.0)
            plt.show(fig)
            
            print("\n\n_____________________________________________\n\n")
    
            input("Press enter to see the distribution of residuals")
            resid=output.resid
            sns.histplot(resid)
            plt.title("Residual Error")
            plt.show(fig)
            print("\n\n_____________________________________________\n\n")
    
            input("\n\nPress enter to see results from the Breusch-Pagan Lagrangian multiplier test for heteroskedascticity ")
            name = ['Lagrange multiplier statistic', 'p-value', 'f-value', 'f p-value']
            test = sm.stats.het_breuschpagan(output.resid, output.model.exog)
            print(lzip(name, test))
            
            input("\n\n*********Press enter to go back to the regression menu*********")
        
        if options=='3':
            print("\n\n_____________________________________________\n\n")
            print(ws['I10'].value)
            input("\n\n*********Press enter to go back to the regression menu*********")
    
        print("\n\n_____________________________________________\n\n")
    
        options=input(ws['I9'].value)
        while options not in ['1','2','3','4']:
                options=input("You have to enter a value from 1 to 4")
    print("\n\n_____________________________________________\n\n")
    
    input("Moving on, in light of those diagnostics we should fit the model with a square term for risk scores. \n\n Press enter to do that.")
    
    ###purge outliers
    
    
    influence = output.get_influence()
    rawcooks= influence.cooks_distance
    Patients['cooks']=rawcooks[0]
    #Patients['cooks'].astype()
    Patients.loc[Patients['cooks']>.016]=np.nan
    del [influence,rawcooks]
    
    model=smf.ols('Annual_med_costs~Risk_pct+I(Risk_pct**2)+Male+SNAP_yes',data=Patients)
    result=model.fit()
    print(result.summary())
    
    print("\n\n_____________________________________________\n\n")
    
    #say we need to graph it
    input(ws['I11'].value)
    
    
    model=smf.ols('Annual_med_costs~Risk_pct+I(Risk_pct**2)',data=Patients)
    result=model.fit()
    
    fig = sm.graphics.plot_fit(result,'Risk_pct', vlines=False)
    plt.show(fig)
    
    print("\n\n_____________________________________________\n\n")
    
    #Conclude the section.
    input(ws['I12'].value)
    
    
    
