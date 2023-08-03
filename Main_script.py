# -*- coding: utf-8 -*-
"""
Created on Fri Jul 28 13:04:50 2023
##Committed to Git 8_3_2023
@author: Matt0
"""
import sys
import Scorers as scorer
import run_analyses as analyze
import pandas as pd
import Weight_selector as ws
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
import Risk_assign as risk
from scipy.stats import chi2_contingency
from openpyxl.workbook import workbook
from openpyxl import load_workbook
import statsmodels.api as sm
from scipy import stats

from statsmodels.compat import lzip
import statsmodels.formula.api as smf
import statsmodels.stats.api as sms
from tabulate import tabulate
sns.set_theme()


#####INITIALIZING

#import the data
Patients=pd.read_csv(r"C:\Users\Matt0\test project\Framingham score tool\Framingham practice more.txt").rename(
    columns={'Total_Cholesterol':'Cholest','HDL_Cholesterol':'HDL',
             'Smoking_Status':'Smoke', 'Systolic_Blood_Pressure':'Systolic',
             'Treatment_for_High_BP':'Treat'})
#purge missings if any
Patients.dropna()

#load the work book with the script
wb=load_workbook("C:/Users/Matt0/test project/Framingham score tool/Script.xlsx")
ws=wb.active

##quieetly prepare the data
Patients,rawscores=analyze.calc_all(Patients)
Patients=pd.concat([Patients,risk.risk_assign(Patients['Framingham'],Patients['Gender'])],axis=1)
Patients=analyze.gendata(Patients)


#Show the welcome message
#Make me a input later
print(ws["A2"].value)
     
demoselect=input(ws['A3'].value)

while demoselect not in["1","2","3","4","5", "6"]:
    demoselect=input("you can only enter values from 1 through 6 here")

while demoselect!="6":
    
    
    
    if demoselect=="1":
    #####Demonstrate the score calculator
    
        analyze.demonstrate()
    
    
        print("\n\n_____________________________________________\n\n")
    
    
    
    if demoselect=="2":
        #####Let you enter your own
        keepgoing="Y"
        
        while keepgoing=="Y":
            print("\n\n_____________________________________________\n\n")
            analyze.enter_new()
            keepgoing=input("\n Would you like to enter another case? Enter Y for yes, N for no")
        

        print("***************** \n\n")

        print("\n\n_____________________________________________\n\n")
    
    if demoselect=="3":
        ####Show high risk patients
        analyze.show_risky(Patients)
    
        print("\n\n_____________________________________________\n\n")
    
    if demoselect=="4":
        ####Run a cross tabulation w/ snap and risk scores

        analyze.crosstab(Patients)
        
        #################
              
    
        print("\n\n_____________________________________________\n\n")
    
    ###press enter to continue###
    ####REGRESSION ANALYSIS of spending
    if demoselect=="5":
        
        #make the histograms
        analyze.makehistogram(Patients)
                
        print("\n\n_____________________________________________\n\n")
        ####Do the scatter plot
        analyze.runscatter(Patients)
        
        
        
        ####Do the continuos model demo
        analyze.OLSdemo(Patients)
    
    #EXIT THE PROGRAM
    if demoselect=="6":
        sys.exit()
        
###BACK TO THE MENU        
    demoselect=input(ws['A3'].value)
    while demoselect not in["1","2","3","4","5", "6"]:
        demoselect=input("you can only enter values from 1 through 6 here")



